from flask import Flask, render_template, request, redirect, url_for, send_file
import json, os, datetime, base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

app = Flask(__name__)

DATA_FILE = "registrations.json"
ADMIN_PASSWORD = "MEINADMINPASSWORT"

# ---------------- JSON ----------------
def load_data():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def generate_id():
    data = load_data()
    return max([r["ID"] for r in data], default=0) + 1

def calculate_age(birth):
    try:
        b = datetime.datetime.strptime(birth, "%Y-%m-%d").date()
        today = datetime.date.today()
        return today.year - b.year - ((today.month, today.day) < (b.month, b.day))
    except:
        return ""
        
# --- Alter & Gruppe berechnen ---
def calculate_age_and_group(birth):
    try:
        b = datetime.datetime.strptime(birth, "%Y-%m-%d").date()
        today = datetime.date.today()
        age = today.year - birthdate.year - (
            (today.month, today.day) < (birthdate.month, birthdate.day)
        )
        if 5 <= age <= 7:
            group = "5-7"
        elif 8 <= age <= 13:
            group = "8-13"
        else:
            group = "?"
        return age, group
    except:
        return None, "?"

# ---------------- ROUTES ----------------
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        data = load_data()
        reg_id = generate_id()

        entry = {
            "ID": reg_id,
            "Geschlecht": request.form.get("Geschlecht"),
            "Vorname": request.form.get("Vorname"),
            "Nachname": request.form.get("Nachname"),
            "Strasse": request.form.get("Strasse"),
            "PLZ": request.form.get("PLZ"),
            "Ort": request.form.get("Ort"),
            "Geburtsdatum": request.form.get("Geburtsdatum"),
            "Notfallnummer": request.form.get("Notfallnummer"),
            "Unterschrift": request.form.get("Unterschrift"),
            "DSGVO": True
        }

        # Initiale Tages-/Verswerte
        for i in range(1, 6):
            entry[f"Tag{i}"] = False
            entry[f"Verse{i}"] = False

        
        
        # Nach Update der Geburtsdatum
        age, group = calculate_age_and_group(r.get("Geburtsdatum", ""))
        r["Alter"] = age
        r["Gruppe"] = group
        
        data.append(entry)
        save_data(data)
        return redirect(url_for("success", reg_id=reg_id))

    return render_template("index.html")

@app.route("/success")
def success():
    return render_template("success.html", reg_id=request.args.get("reg_id"))

# --- Admin ---
@app.route("/admin", methods=["GET"])
def admin():
    pw = request.args.get("pw")
    if pw != ADMIN_PASSWORD:
        return "Zugriff verweigert", 403

    data = load_data()
    
    for r in data:
        age, group = calculate_age_and_group(r.get("Geburtsdatum", ""))
        r["Alter"] = age
        r["Gruppe"] = group
    
    search = request.args.get("search", "").lower()
    day_filter = request.args.get("day", "")

    if search:
        data = [
            r for r in data
            if search in str(r["ID"])
            or search in r["Vorname"].lower()
            or search in r["Nachname"].lower()
        ]

    if day_filter:
        data = [r for r in data if r.get(day_filter)]

    return render_template(
        "admin.html",
        registrations=data,
        pw=pw,
        search=search,
        day_filter=day_filter
    )
    
# --- Update Eintrag ---
@app.route("/update/<int:reg_id>", methods=["POST"])
def update_entry(reg_id):
    pw = request.args.get("pw")
    if pw != ADMIN_PASSWORD:
        return "Zugriff verweigert", 403

    data = load_data()
    for r in data:
        if r["ID"] == reg_id:
            for key in r:
                if key.startswith("Tag") or key.startswith("Verse"):
                    r[key] = key in request.form
                elif key in request.form:
                    r[key] = request.form.get(key)

            r["Punkte"] = sum(
                r[f"Tag{i}"] + r[f"Verse{i}"] for i in range(1, 6)
            )
            break

    save_data(data)
    return redirect(url_for("admin", pw=pw))

@app.route("/delete/<int:reg_id>")
def delete_entry(reg_id):
    pw = request.args.get("pw")
    if pw != ADMIN_PASSWORD:
        return "Zugriff verweigert", 403

    data = [r for r in load_data() if r["ID"] != reg_id]
    save_data(data)
    return redirect(url_for("admin", pw=pw))

@app.route("/export_excel")
def export_excel():
    pw = request.args.get("pw")
    if pw != ADMIN_PASSWORD:
        return "Zugriff verweigert", 403

    data = load_data()
    wb = Workbook()
    ws = wb.active

    headers = list(data[0].keys())
    ws.append(headers)

    row = 2
    for r in data:
        col = 1
        for h in headers:
            if h == "Unterschrift" and r[h].startswith("data:image"):
                img = XLImage(BytesIO(base64.b64decode(r[h].split(",")[1])))
                ws.add_image(img, ws.cell(row=row, column=col).coordinate)
            else:
                ws.cell(row=row, column=col, value=r[h])
            col += 1
        row += 1

    wb.save("export.xlsx")
    return send_file("export.xlsx", as_attachment=True)

@app.route("/datenschutz")
def datenschutz():
    return render_template("datenschutz.html")
